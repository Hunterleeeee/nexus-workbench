"""工作项 / 产物 / 关系通用数据层领域。

拆自 app.py（2026-08-14 第二十批）。包含: work_items/artifacts/relations 数据层与路由、
文档内容提取（mineru/markitdown）、safe_filename 等通用工具。
仍在 app.py 的领域函数经 _app_call 运行时转发。
"""
from __future__ import annotations

import io
import json
import os
import re
import shutil
import sqlite3
import subprocess
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from fastapi import HTTPException
from pydantic import BaseModel, Field

from .core import (
    DATA_DIR,
    KNOWLEDGE_DIR,
    OBSIDIAN_VAULT_DIR,
    OUTPUTS_DIR,
    ROOT,
    clip,
    decode_json_column,
    log,
    now_iso,
    save_json_atomic,
)
from .db import db_connection
from .instance import app
from .agent_platform import AGENT_REGISTRY
from .browser import work_item_source_context
from .notifications import create_notification_record
from .projects import PROJECT_LINKS, agent_display_name, project_audit, project_link_summary, public_project_link


def _app_call(fn_name: str, *args: Any, **kwargs: Any) -> Any:
    """通过 app 命名空间调用仍在 app.py 的领域函数——测试 patch app.X 时能生效。"""
    import app as _app

    return getattr(_app, fn_name)(*args, **kwargs)


class WorkItemRequest(BaseModel):
    title: str = Field(min_length=1, max_length=240)
    description: str = Field(default="", max_length=20_000)
    kind: str = Field(default="task", max_length=50)
    status: str = Field(default="open", max_length=30)
    priority: str = Field(default="normal", max_length=20)
    source_project: str = Field(default="workbench", max_length=80)
    target_project: str = Field(default="", max_length=80)
    metadata: dict[str, Any] = Field(default_factory=dict)


class WorkItemUpdateRequest(BaseModel):
    status: str | None = Field(default=None, max_length=30)
    priority: str | None = Field(default=None, max_length=20)
    description: str | None = Field(default=None, max_length=20_000)
    target_project: str | None = Field(default=None, max_length=80)
    metadata: dict[str, Any] | None = None


class HandoffRequest(BaseModel):
    title: str = Field(min_length=1, max_length=240)
    description: str = Field(default="", max_length=20_000)
    from_project: str = Field(default="workbench", max_length=80)
    to_project: str = Field(min_length=1, max_length=80)
    priority: str = Field(default="normal", max_length=20)
    metadata: dict[str, Any] = Field(default_factory=dict)
    confirmed: bool = False


class ArtifactRequest(BaseModel):
    project_id: str = Field(min_length=1, max_length=80)
    name: str = Field(min_length=1, max_length=240)
    path: str = Field(default="", max_length=1000)
    kind: str = Field(default="file", max_length=50)
    metadata: dict[str, Any] = Field(default_factory=dict)

def work_item_next_step_quality(item: dict[str, Any]) -> dict[str, Any]:
    """Describe whether a WorkItem contains an actionable, reviewable next step.

    This is intentionally deterministic and conservative.  It does not infer an
    action from a long description: only explicitly captured ``next_steps`` (or
    ``next_step``) count as a next step.  A target, owner, or due date makes the
    step ready to route; otherwise it remains reviewable instead of being shown
    as executable.
    """
    metadata = item.get("metadata") if isinstance(item.get("metadata"), dict) else {}
    raw_steps = metadata.get("next_steps")
    if raw_steps is None:
        raw_steps = metadata.get("next_step")
    if isinstance(raw_steps, str):
        raw_steps = re.split(r"[\n；;]+", raw_steps)
    if not isinstance(raw_steps, list):
        raw_steps = []
    steps = [clip(str(step).strip(), 500) for step in raw_steps if str(step).strip()][:5]
    target = str(item.get("target_project") or metadata.get("target_project") or "").strip()
    owner = str(metadata.get("owner") or metadata.get("assignee") or metadata.get("next_step_owner") or "").strip()
    due_at = str(item.get("due_at") or metadata.get("due_at") or "").strip()
    source = str(metadata.get("next_steps_source") or metadata.get("next_step_source") or "").strip()
    if not steps:
        status = "missing"
        label = "需补下一步"
        next_action = "补一条最小可执行动作；不要只保留背景描述。"
    elif target or owner or due_at:
        status = "ready"
        label = "下一步清楚"
        next_action = "可以领取或执行；执行前仍按项目权限确认外部动作。"
    else:
        status = "review"
        label = "需确认范围"
        next_action = "补充目标 Agent、负责人或截止时间，再进入主动协作队列。"
    return {
        "status": status,
        "label": label,
        "steps": steps,
        "source": source or "未记录",
        "target_project": target,
        "owner": owner,
        "due_at": due_at,
        "next_action": next_action,
    }


def work_item_row(row: sqlite3.Row) -> dict[str, Any]:
    item = {key: row[key] for key in row.keys()}
    item["metadata"] = decode_json_column(item.pop("metadata_json", "{}"))
    item["result"] = decode_json_column(item.pop("result_json", "{}"))
    item["source_context"] = work_item_source_context(item["metadata"])
    source_project = item.get("source_project", "workbench")
    target_projects = [project_id.strip() for project_id in str(item.get("target_project", "")).split(",") if project_id.strip()]
    item["source_agent_name"] = agent_display_name(source_project)
    item["target_agent_names"] = [agent_display_name(project_id) for project_id in target_projects]
    item["target_agent_label"] = "、".join(item["target_agent_names"])
    item["claimed"] = bool(item.get("claimed_at"))
    item["next_step_quality"] = _app_call('work_item_next_step_quality', item)
    return item


def relation_row(row: sqlite3.Row) -> dict[str, Any]:
    relation = {key: row[key] for key in row.keys()}
    relation["metadata"] = decode_json_column(relation.pop("metadata_json", "{}"))
    return relation


def artifact_row(row: sqlite3.Row) -> dict[str, Any]:
    artifact = {key: row[key] for key in row.keys()}
    artifact["metadata"] = decode_json_column(artifact.pop("metadata_json", "{}"))
    return artifact


def list_artifacts(project_id: str = "") -> list[dict[str, Any]]:
    connection = db_connection()
    try:
        if project_id:
            rows = connection.execute(
                "SELECT * FROM artifacts WHERE project_id = ? ORDER BY created_at DESC LIMIT 200", (project_id,)
            ).fetchall()
        else:
            rows = connection.execute("SELECT * FROM artifacts ORDER BY created_at DESC LIMIT 200").fetchall()
        return [_app_call('artifact_row', row) for row in rows]
    finally:
        connection.close()


def get_artifact_record(artifact_id: int) -> dict[str, Any] | None:
    connection = db_connection()
    try:
        row = connection.execute("SELECT * FROM artifacts WHERE id = ?", (int(artifact_id),)).fetchone()
        return _app_call('artifact_row', row) if row else None
    finally:
        connection.close()


DOCUMENT_FACTORY_SOURCE_SUFFIXES = {".txt", ".md", ".markdown", ".csv", ".json", ".pdf", ".docx", ".xlsx", ".xlsm", ".pptx", ".html", ".htm"}


def document_factory_allowed_roots() -> list[Path]:
    """Only expose files already inside Workbench-managed material roots.

    Artifact metadata can come from multiple projects. The document Agent may
    read their registered files, but it must not turn an arbitrary database
    row into an unrestricted filesystem reader.
    """
    roots: list[Path] = []
    for root in (OUTPUTS_DIR, KNOWLEDGE_DIR, OBSIDIAN_VAULT_DIR, DATA_DIR):
        try:
            resolved = root.expanduser().resolve()
        except OSError:
            continue
        if resolved not in roots:
            roots.append(resolved)
    return roots


def artifact_source_path(artifact: dict[str, Any]) -> tuple[Path | None, str]:
    raw_path = str(artifact.get("path") or "").strip()
    if not raw_path:
        return None, "这个 Artifact 没有登记文件路径"
    try:
        candidate = Path(raw_path).expanduser()
        if not candidate.is_absolute():
            candidate = ROOT / candidate
        resolved = candidate.resolve()
    except OSError:
        return None, "Artifact 文件路径无法解析"
    if resolved.suffix.lower() not in DOCUMENT_FACTORY_SOURCE_SUFFIXES:
        return None, f"暂不支持读取 {resolved.suffix or '该'} 文件格式"
    data_root = DATA_DIR.expanduser().resolve()
    is_data_file = resolved == data_root or data_root in resolved.parents
    if is_data_file and str(artifact.get("project_id") or "") not in {"aihot", "market", "server"}:
        return None, "账户和系统配置快照可能包含敏感字段，不作为文档材料读取"
    if not any(resolved == root or root in resolved.parents for root in _app_call('document_factory_allowed_roots', )):
        return None, "出于安全边界，只能读取工作台已管理目录中的文件"
    if not resolved.is_file():
        return None, "登记的文件已经不存在"
    return resolved, ""


MINERU_COMMAND = os.getenv("WORKBENCH_MINERU_CMD", "mineru").strip() or "mineru"
MINERU_TIMEOUT_SECONDS = max(60, int(os.getenv("WORKBENCH_MINERU_TIMEOUT_SECONDS", "600") or 600))
MINERU_SUFFIXES = {".pdf", ".png", ".jpg", ".jpeg", ".webp"}


def mineru_status() -> dict[str, Any]:
    """Report the optional MinerU adapter.

    MinerU is markedly better than pypdf/MarkItDown on Chinese PDFs, scanned
    documents, formulas and complex tables, but it is heavy and slow. It stays
    strictly optional: when the binary is absent the existing chain is used
    unchanged.
    """
    executable = shutil.which(MINERU_COMMAND)
    if not executable:
        return {"available": False, "label": "MinerU 未安装", "version": "", "mode": "off"}
    try:
        probe = subprocess.run([executable, "--version"], capture_output=True, text=True, timeout=20, check=False)
        version = (probe.stdout or probe.stderr or "").strip().splitlines()[0] if probe.returncode == 0 else "已安装"
    except (OSError, subprocess.SubprocessError):
        version = "已安装"
    return {"available": True, "label": "MinerU 可用", "version": clip(version, 60), "mode": "preferred", "path": executable}


def extract_with_mineru(raw: bytes, filename: str) -> str:
    """Run MinerU in a temp dir and return the Markdown it produced.

    Returns "" on any failure so the caller falls through to MarkItDown and
    then the native parsers. MinerU runs as a subprocess rather than an import
    so a heavy optional dependency never lives inside the API process, and a
    hung run cannot outlive its timeout.
    """
    suffix = Path(filename).suffix.lower()
    if suffix not in MINERU_SUFFIXES:
        return ""
    executable = shutil.which(MINERU_COMMAND)
    if not executable:
        return ""
    import tempfile

    try:
        with tempfile.TemporaryDirectory(prefix="workbench-mineru-") as workspace:
            root = Path(workspace)
            source = root / f"input{suffix}"
            source.write_bytes(raw)
            output = root / "out"
            output.mkdir(parents=True, exist_ok=True)
            completed = subprocess.run(
                [executable, "-p", str(source), "-o", str(output)],
                capture_output=True,
                text=True,
                timeout=MINERU_TIMEOUT_SECONDS,
                check=False,
                start_new_session=True,
            )
            if completed.returncode != 0:
                return ""
            markdowns = sorted(output.rglob("*.md"), key=lambda path: path.stat().st_size, reverse=True)
            for candidate in markdowns:
                try:
                    text = candidate.read_text(encoding="utf-8").strip()
                except (OSError, UnicodeDecodeError):
                    continue
                if text:
                    return text
    except (OSError, subprocess.SubprocessError, ValueError):
        return ""
    return ""


def markitdown_status() -> dict[str, Any]:
    """Report the optional Microsoft MarkItDown adapter without making it required.

    MarkItDown improves table, slide and mixed-document extraction when it is
    installed. The native parsers remain the deterministic fallback so a
    missing optional package never blocks the document factory.
    """
    try:
        import markitdown  # type: ignore

        version = str(getattr(markitdown, "__version__", "") or "已安装")
        return {"available": True, "label": "MarkItDown 可用", "version": version, "mode": "optional"}
    except ImportError:
        return {"available": False, "label": "内置解析器", "version": "", "mode": "fallback"}
    except Exception as exc:
        return {"available": False, "label": "内置解析器", "version": "", "mode": "fallback", "detail": clip(str(exc), 120)}


def extract_with_markitdown(raw: bytes, filename: str) -> str:
    """Try MarkItDown for rich office/PDF/HTML extraction, returning empty on fallback."""
    suffix = Path(filename).suffix.lower()
    if suffix in {".txt", ".md", ".markdown", ".csv", ".json"}:
        return ""
    try:
        from markitdown import MarkItDown  # type: ignore
    except ImportError:
        return ""
    try:
        converter = MarkItDown()
        stream = io.BytesIO(raw)
        result = None
        try:
            from markitdown import StreamInfo  # type: ignore

            result = converter.convert_stream(stream, stream_info=StreamInfo(file_extension=suffix))
        except (ImportError, TypeError):
            stream.seek(0)
            result = converter.convert_stream(stream)
        text = str(getattr(result, "text_content", "") or "").strip()
        return text
    except Exception:
        # The native parser below is intentionally the compatibility path for
        # older MarkItDown releases and files that the optional adapter cannot
        # decode.
        return ""


def document_extraction_engine(filename: str) -> str:
    suffix = Path(filename).suffix.lower()
    if suffix in {".txt", ".md", ".markdown", ".csv", ".json"}:
        return "内置文本解析器"
    if suffix in MINERU_SUFFIXES and _app_call('mineru_status', )["available"]:
        return "MinerU（优先）→ MarkItDown → 内置解析器"
    return "MarkItDown（可选）或内置解析器"


def extract_document_bytes(raw: bytes, filename: str) -> str:
    suffix = Path(filename).suffix.lower()
    if suffix in {".txt", ".md", ".markdown", ".csv", ".json"}:
        return raw.decode("utf-8", errors="replace")
    # MinerU first for PDFs and images: it handles Chinese layout, scans,
    # formulas and merged tables that the other parsers silently mangle.
    # Everything below stays as the fallback chain.
    mineru_text = _app_call('extract_with_mineru', raw, filename)
    if mineru_text:
        return mineru_text
    enhanced = _app_call('extract_with_markitdown', raw, filename)
    if enhanced:
        return enhanced
    if suffix == ".pdf":
        from pypdf import PdfReader

        reader = PdfReader(io.BytesIO(raw))
        return "\n\n".join(page.extract_text() or "" for page in reader.pages)
    if suffix == ".docx":
        from docx import Document

        document = Document(io.BytesIO(raw))
        return "\n\n".join(paragraph.text for paragraph in document.paragraphs)
    if suffix in {".xlsx", ".xlsm"}:
        from openpyxl import load_workbook

        workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        sheets = []
        for sheet in workbook.worksheets:
            rows = [" | ".join("" if value is None else str(value) for value in row) for row in sheet.iter_rows(values_only=True)]
            sheets.append(f"## {sheet.title}\n" + "\n".join(rows))
        return "\n\n".join(sheets)
    if suffix in {".pptx", ".html", ".htm"}:
        raise ValueError("这类文件需要安装可选的 MarkItDown 才能读取；也可以先导出为 PDF 或 Markdown")
    raise ValueError("暂不支持该文件格式，请上传 Markdown、TXT、CSV、JSON、PDF、DOCX、XLSX 或 PPTX")


def read_artifact_source(artifact: dict[str, Any]) -> tuple[str, str]:
    path, error = _app_call('artifact_source_path', artifact)
    if error or not path:
        return "", error or "Artifact 文件不可读"
    try:
        if path.stat().st_size > 15 * 1024 * 1024:
            return "", "文件超过 15 MB，暂不作为文档材料读取"
        content = _app_call('extract_document_bytes', path.read_bytes(), path.name).strip()
    except Exception as exc:
        return "", f"读取失败：{clip(str(exc), 180)}"
    if not content:
        return "", "文件中没有可提取的文本"
    return clip(content, 100_000), ""


def create_artifact_record(
    *, project_id: str, name: str, path: str = "", kind: str = "file", metadata: dict[str, Any] | None = None
) -> dict[str, Any]:
    connection = db_connection()
    try:
        cursor = connection.execute(
            """INSERT INTO artifacts
            (project_id, name, path, kind, metadata_json, created_at)
            VALUES (?, ?, ?, ?, ?, ?)""",
            (project_id, name.strip(), path.strip(), kind.strip() or "file", json.dumps(metadata or {}, ensure_ascii=False), now_iso()),
        )
        connection.commit()
        row = connection.execute("SELECT * FROM artifacts WHERE id = ?", (cursor.lastrowid,)).fetchone()
        return _app_call('artifact_row', row)
    finally:
        connection.close()


def register_artifact_safely(
    *, project_id: str, name: str, path: str = "", kind: str = "file", metadata: dict[str, Any] | None = None
) -> dict[str, Any] | None:
    try:
        return _app_call('create_artifact_record', project_id=project_id, name=name, path=path, kind=kind, metadata=metadata)
    except Exception:
        # Artifact indexing must never make the primary project operation fail.
        return None


def list_work_items(status: str = "all", project_id: str = "") -> list[dict[str, Any]]:
    connection = db_connection()
    try:
        clauses: list[str] = []
        values: list[str] = []
        if status and status != "all":
            clauses.append("status = ?")
            values.append(status)
        if project_id:
            clauses.append("(source_project = ? OR target_project = ?)")
            values.extend([project_id, project_id])
        where = f" WHERE {' AND '.join(clauses)}" if clauses else ""
        rows = connection.execute(
            f"SELECT * FROM work_items{where} ORDER BY updated_at DESC, id DESC LIMIT 200", values
        ).fetchall()
        return [_app_call('work_item_row', row) for row in rows]
    finally:
        connection.close()


def get_work_item_record(item_id: int) -> dict[str, Any] | None:
    connection = db_connection()
    try:
        row = connection.execute("SELECT * FROM work_items WHERE id = ?", (item_id,)).fetchone()
        return _app_call('work_item_row', row) if row else None
    finally:
        connection.close()


def create_work_item_record(
    *,
    title: str,
    description: str = "",
    kind: str = "task",
    status: str = "open",
    priority: str = "normal",
    source_project: str = "workbench",
    target_project: str = "",
    metadata: dict[str, Any] | None = None,
) -> dict[str, Any]:
    timestamp = now_iso()
    connection = db_connection()
    try:
        cursor = connection.execute(
            """INSERT INTO work_items
            (title, description, kind, status, priority, source_project, target_project, metadata_json, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                title.strip(),
                description.strip(),
                kind.strip() or "task",
                status.strip() or "open",
                priority.strip() or "normal",
                source_project.strip() or "workbench",
                target_project.strip(),
                json.dumps(metadata or {}, ensure_ascii=False),
                timestamp,
                timestamp,
            ),
        )
        connection.commit()
        row = connection.execute("SELECT * FROM work_items WHERE id = ?", (cursor.lastrowid,)).fetchone()
        item = _app_call('work_item_row', row)
    finally:
        connection.close()
    # Agent dispatches get one result notification after the orchestrator finishes.
    # Emitting a second "task created" notification here made the in-app center
    # noisy and left the user without the actual result.
    if kind != "agent_dispatch" and (kind == "alert" or priority in {"high", "urgent"}):
        try:
            item_metadata = item.get("metadata") if isinstance(item.get("metadata"), dict) else {}
            notification_project = item_metadata.get("notification_project") or target_project or source_project or "workbench"
            create_notification_record(
                title=item["title"],
                body=item.get("description", ""),
                project_id=notification_project,
                kind=kind,
                level="critical" if priority == "urgent" else "warning" if priority == "high" else "info",
                href="/" if notification_project == "workbench" else f"/projects/{notification_project}",
                event_key=f"work-item:{item['id']}",
                dedupe_seconds=0,
            )
        except Exception:
            # Notification delivery must never make creation of the primary work item fail.
            log.debug("忽略异常（create_work_item_record）", exc_info=True)
    return item


def update_work_item_record(item_id: int, values: dict[str, Any]) -> dict[str, Any] | None:
    allowed = {"status", "priority", "description", "target_project", "metadata_json", "claimed_at", "claimed_run_id", "result_json", "completed_at", "last_error"}
    updates = [(key, value) for key, value in values.items() if key in allowed]
    if not updates:
        return next((item for item in _app_call('list_work_items', ) if item["id"] == item_id), None)
    updates.append(("updated_at", now_iso()))
    connection = db_connection()
    try:
        assignments = ", ".join(f"{key} = ?" for key, _ in updates)
        cursor = connection.execute(
            f"UPDATE work_items SET {assignments} WHERE id = ?",
            [value for _, value in updates] + [item_id],
        )
        connection.commit()
        if cursor.rowcount == 0:
            return None
        row = connection.execute("SELECT * FROM work_items WHERE id = ?", (item_id,)).fetchone()
        return _app_call('work_item_row', row)
    finally:
        connection.close()


def create_relation_record(
    *, from_type: str, from_id: str, to_type: str, to_id: str, relation_type: str, metadata: dict[str, Any] | None = None
) -> dict[str, Any]:
    connection = db_connection()
    try:
        cursor = connection.execute(
            """INSERT INTO relations
            (from_type, from_id, to_type, to_id, relation_type, metadata_json, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?)""",
            (from_type, str(from_id), to_type, str(to_id), relation_type, json.dumps(metadata or {}, ensure_ascii=False), now_iso()),
        )
        connection.commit()
        row = connection.execute("SELECT * FROM relations WHERE id = ?", (cursor.lastrowid,)).fetchone()
        return _app_call('relation_row', row)
    finally:
        connection.close()


def list_relations(entity_id: str = "") -> list[dict[str, Any]]:
    connection = db_connection()
    try:
        if entity_id:
            rows = connection.execute(
                "SELECT * FROM relations WHERE from_id = ? OR to_id = ? ORDER BY created_at DESC LIMIT 200",
                (str(entity_id), str(entity_id)),
            ).fetchall()
        else:
            rows = connection.execute("SELECT * FROM relations ORDER BY created_at DESC LIMIT 200").fetchall()
        return [_app_call('relation_row', row) for row in rows]
    finally:
        connection.close()


def safe_filename(value: str, fallback: str = "output") -> str:
    cleaned = re.sub(r"[^\w\-\u4e00-\u9fff]+", "-", value, flags=re.UNICODE).strip("-")
    return (cleaned or fallback)[:80]

@app.get("/api/artifacts")
def get_artifacts(project_id: str = "") -> dict[str, Any]:
    return {"artifacts": _app_call('list_artifacts', project_id)}


@app.post("/api/artifacts")
def create_artifact(request: ArtifactRequest) -> dict[str, Any]:
    if request.project_id not in AGENT_REGISTRY:
        raise HTTPException(400, "来源项目 Agent 不存在")
    payload = request.model_dump() if hasattr(request, "model_dump") else request.dict()
    return {"artifact": _app_call('create_artifact_record', **payload)}


@app.get("/api/relations")
def get_relations(entity_id: str = "") -> dict[str, Any]:
    return {"relations": _app_call('list_relations', entity_id)}


@app.get("/api/project-links")
async def get_project_links(project_id: str = "") -> dict[str, Any]:
    """Return the project handoff graph, optionally narrowed to one project."""
    if project_id:
        summary = project_link_summary(project_id)
        return {"project_id": project_id, "links": [public_project_link(edge) for edge in summary["inbound"] + summary["outbound"]]}
    return {"links": [public_project_link(edge) for edge in PROJECT_LINKS]}


@app.get("/api/project-audit")
def get_project_audit(project_id: str = "") -> dict[str, Any]:
    if project_id and project_id not in AGENT_REGISTRY:
        raise HTTPException(404, "项目 Agent 不存在")
    return project_audit(project_id)


@app.get("/api/work-items")
def get_work_items(status: str = "all", project_id: str = "") -> dict[str, Any]:
    return {"items": _app_call('list_work_items', status=status, project_id=project_id)}


@app.post("/api/work-items")
def create_work_item(request: WorkItemRequest) -> dict[str, Any]:
    if request.status not in {"open", "running", "blocked", "done", "archived", "failed"}:
        raise HTTPException(400, "不支持的工作项状态")
    payload = request.model_dump() if hasattr(request, "model_dump") else request.dict()
    return {"item": _app_call('create_work_item_record', **payload)}


@app.patch("/api/work-items/{item_id}")
def update_work_item(item_id: int, request: WorkItemUpdateRequest) -> dict[str, Any]:
    values = request.model_dump(exclude_unset=True) if hasattr(request, "model_dump") else request.dict(exclude_unset=True)
    if values.get("status") and values["status"] not in {"open", "running", "blocked", "done", "archived", "failed"}:
        raise HTTPException(400, "不支持的工作项状态")
    if "metadata" in values:
        values["metadata_json"] = json.dumps(values.pop("metadata") or {}, ensure_ascii=False)
    item = _app_call('update_work_item_record', item_id, values)
    if not item:
        raise HTTPException(404, "工作项不存在")
    return {"item": item}


__all__ = [
    "WorkItemRequest",
    "WorkItemUpdateRequest",
    "HandoffRequest",
    "ArtifactRequest",
    "work_item_next_step_quality",
    "work_item_row",
    "relation_row",
    "artifact_row",
    "list_artifacts",
    "get_artifact_record",
    "DOCUMENT_FACTORY_SOURCE_SUFFIXES",
    "document_factory_allowed_roots",
    "artifact_source_path",
    "MINERU_COMMAND",
    "MINERU_TIMEOUT_SECONDS",
    "MINERU_SUFFIXES",
    "mineru_status",
    "extract_with_mineru",
    "markitdown_status",
    "extract_with_markitdown",
    "document_extraction_engine",
    "extract_document_bytes",
    "read_artifact_source",
    "create_artifact_record",
    "register_artifact_safely",
    "list_work_items",
    "get_work_item_record",
    "create_work_item_record",
    "update_work_item_record",
    "create_relation_record",
    "list_relations",
    "safe_filename",
    "get_artifacts",
    "create_artifact",
    "get_relations",
    "get_project_links",
    "get_project_audit",
    "get_work_items",
    "create_work_item",
    "update_work_item",
]
